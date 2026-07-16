"""Normalize the authoritative analysis-factor workbook into backend JSON."""

from __future__ import annotations

import argparse
import hashlib
import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


CORE_SHEETS = {
    "common parameters",
    "Qualitative Guide",
    "sector names",
    "Qualitative Framework",
    "Super Master Prompt",
}


def _json_value(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _rows(sheet) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for row in sheet.iter_rows():
        cells = [_json_value(cell.value) for cell in row]
        while cells and cells[-1] is None:
            cells.pop()
        if any(value is not None for value in cells):
            output.append({"row": row[0].row, "cells": cells})
    return output


def _table(sheet, header_row: int, first_column: int) -> list[dict[str, Any]]:
    headers = [cell.value for cell in sheet[header_row]][first_column - 1 :]
    while headers and headers[-1] is None:
        headers.pop()
    normalized_headers = [str(value).strip() for value in headers]
    output = []
    for row in sheet.iter_rows(min_row=header_row + 1):
        values = [_json_value(cell.value) for cell in row][first_column - 1 : first_column - 1 + len(normalized_headers)]
        if not any(value is not None for value in values):
            continue
        output.append(dict(zip(normalized_headers, values)))
    return output


def build_registry(source: Path) -> dict[str, Any]:
    workbook = load_workbook(source, data_only=False, read_only=False)
    digest = hashlib.sha256(source.read_bytes()).hexdigest()
    return {
        "schema_version": 1,
        "source": {
            "filename": source.name,
            "sha256": digest,
            "verification_status": "user_verified_authoritative",
        },
        "common_parameters": _table(workbook["common parameters"], 1, 2),
        "sector_map": _table(workbook["sector names"], 2, 2),
        "qualitative_guide": _rows(workbook["Qualitative Guide"]),
        "qualitative_framework": _rows(workbook["Qualitative Framework"]),
        "sector_factors": {
            sheet.title: _rows(sheet)
            for sheet in workbook.worksheets
            if sheet.title not in CORE_SHEETS
        },
        "workflow_reference": _rows(workbook["Super Master Prompt"]),
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    registry = build_registry(args.source)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(registry, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(json.dumps({
        "output": str(args.output),
        "sha256": registry["source"]["sha256"],
        "common_parameters": len(registry["common_parameters"]),
        "sectors": len(registry["sector_factors"]),
    }))


if __name__ == "__main__":
    main()

