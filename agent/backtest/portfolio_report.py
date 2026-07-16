"""Numeric-first portfolio workbook produced after every equity backtest.

The workbook is an audit/report artifact.  Tax rates are deliberately left as
manual inputs and are applied only after the pre-tax simulation has completed.
"""

from __future__ import annotations

import math
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill


_HEADER_FILL = PatternFill("solid", fgColor="17324D")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_PERCENT_FORMAT = "0.00%"
_NUMBER_FORMAT = "#,##0.00"


def _number(value: Any) -> float | int | None:
    """Return a finite Excel-safe number without converting labels to text."""
    if isinstance(value, bool):
        return int(value)
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(number):
        return None
    return int(number) if number.is_integer() else number


def _style_table(sheet, *, freeze: str = "A2") -> None:
    sheet.freeze_panes = freeze
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    for column in sheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 28)
        sheet.column_dimensions[column[0].column_letter].width = max(width, 11)


def _append_frame(sheet, frame: pd.DataFrame) -> None:
    sheet.append([str(column) for column in frame.columns])
    for row in frame.itertuples(index=False, name=None):
        sheet.append([
            value.to_pydatetime() if isinstance(value, pd.Timestamp) else _number(value) if not isinstance(value, str) else value
            for value in row
        ])
    _style_table(sheet)


def _annual_frame(equity: pd.DataFrame) -> pd.DataFrame:
    indexed = equity.copy()
    indexed["timestamp"] = pd.to_datetime(indexed["timestamp"], errors="coerce")
    indexed = indexed.dropna(subset=["timestamp"]).sort_values("timestamp")
    rows: list[dict[str, float | int]] = []
    for year, group in indexed.groupby(indexed["timestamp"].dt.year):
        portfolio_start = float(group["equity"].iloc[0])
        portfolio_end = float(group["equity"].iloc[-1])
        benchmark_start = float(group["benchmark_equity"].iloc[0])
        benchmark_end = float(group["benchmark_equity"].iloc[-1])
        portfolio_return = portfolio_end / portfolio_start - 1 if portfolio_start else 0.0
        benchmark_return = benchmark_end / benchmark_start - 1 if benchmark_start else 0.0
        daily_returns = pd.to_numeric(group["ret"], errors="coerce").dropna()
        rows.append({
            "year": int(year),
            "start_equity": portfolio_start,
            "end_equity": portfolio_end,
            "portfolio_return": portfolio_return,
            "benchmark_return": benchmark_return,
            "excess_return": portfolio_return - benchmark_return,
            "max_drawdown": float(pd.to_numeric(group["drawdown"], errors="coerce").min()),
            "volatility": float(daily_returns.std(ddof=0) * math.sqrt(252)) if len(daily_returns) else 0.0,
        })
    return pd.DataFrame(rows)


def write_portfolio_workbook(
    path: Path,
    *,
    equity: pd.DataFrame,
    positions: pd.DataFrame,
    trades: pd.DataFrame,
    metrics: dict[str, Any],
    master_factors: list[dict[str, Any]] | None = None,
    sector_factors: list[dict[str, Any]] | None = None,
    factor_authority: dict[str, Any] | None = None,
    price_history: pd.DataFrame | None = None,
    ltcg_holding_days: int = 365,
) -> Path:
    """Write numeric history, benchmark, drawdown, holdings and tax sheets."""
    path.parent.mkdir(parents=True, exist_ok=True)
    equity_frame = equity.reset_index() if "timestamp" not in equity.columns else equity.copy()
    positions_frame = positions.reset_index() if "timestamp" not in positions.columns else positions.copy()
    annual = _annual_frame(equity_frame)

    wb = Workbook()
    wb.remove(wb.active)

    summary = wb.create_sheet("Summary")
    summary.append(["metric", "value"])
    for key, value in metrics.items():
        numeric = _number(value)
        if numeric is not None:
            summary.append([str(key), numeric])
    _style_table(summary)

    master_sheet = wb.create_sheet("Master Factors")
    _append_frame(master_sheet, pd.DataFrame(master_factors or []))

    sector_sheet = wb.create_sheet("Sector Factors")
    raw_sector_rows = sector_factors or []
    max_cells = max((len(item.get("cells", [])) for item in raw_sector_rows), default=0)
    sector_frame = pd.DataFrame(
        [
            {
                "source_row": item.get("row"),
                **{f"field_{index + 1}": value for index, value in enumerate(item.get("cells", []))},
            }
            for item in raw_sector_rows
        ],
        columns=["source_row", *[f"field_{index + 1}" for index in range(max_cells)]],
    )
    _append_frame(sector_sheet, sector_frame)

    annual_sheet = wb.create_sheet("Annual")
    _append_frame(annual_sheet, annual)
    for row in annual_sheet.iter_rows(min_row=2, min_col=4, max_col=8):
        for cell in row:
            cell.number_format = _PERCENT_FORMAT

    equity_sheet = wb.create_sheet("Equity")
    _append_frame(equity_sheet, equity_frame)

    prices = price_history.copy() if price_history is not None else pd.DataFrame()
    if not prices.empty:
        prices = prices.sort_index()
        prices.index.name = "timestamp"
        prices = prices.apply(pd.to_numeric, errors="coerce")
    price_frame = prices.reset_index() if not prices.empty else pd.DataFrame(columns=["timestamp"])
    price_sheet = wb.create_sheet("Price History")
    _append_frame(price_sheet, price_frame)

    indexed = prices.copy()
    for column in indexed.columns:
        valid = indexed[column].dropna()
        indexed[column] = indexed[column] / valid.iloc[0] * 100.0 if not valid.empty and valid.iloc[0] != 0 else float("nan")
    price_index_sheet = wb.create_sheet("Price Index")
    _append_frame(
        price_index_sheet,
        indexed.reset_index() if not indexed.empty else pd.DataFrame(columns=["timestamp"]),
    )
    benchmark = equity_frame[["timestamp", "equity", "benchmark_equity", "active_ret"]].copy()
    benchmark_sheet = wb.create_sheet("Benchmark")
    _append_frame(benchmark_sheet, benchmark)

    drawdown = equity_frame[["timestamp", "drawdown"]].copy()
    drawdown_sheet = wb.create_sheet("Drawdown")
    _append_frame(drawdown_sheet, drawdown)
    for cell in drawdown_sheet["B"][1:]:
        cell.number_format = _PERCENT_FORMAT

    positions_sheet = wb.create_sheet("Holdings")
    _append_frame(positions_sheet, positions_frame)
    for row in positions_sheet.iter_rows(min_row=2, min_col=2):
        for cell in row:
            cell.number_format = _PERCENT_FORMAT

    trade_frame = trades.copy()
    if trade_frame.empty:
        trade_frame = pd.DataFrame(columns=[
            "timestamp", "code", "side", "price", "qty", "reason", "pnl", "holding_days", "return_pct"
        ])
    transactions_sheet = wb.create_sheet("Transactions")
    _append_frame(transactions_sheet, trade_frame)
    exits = trade_frame[pd.to_numeric(trade_frame.get("pnl"), errors="coerce").fillna(0).ne(0)].copy()
    holding_days = pd.to_numeric(exits.get("holding_days"), errors="coerce").fillna(0)
    stcg = exits[holding_days <= ltcg_holding_days].copy()
    ltcg = exits[holding_days > ltcg_holding_days].copy()

    for name, frame in (("STCG", stcg), ("LTCG", ltcg)):
        sheet = wb.create_sheet(name)
        _append_frame(sheet, frame)

    tax = wb.create_sheet("Manual Tax")
    tax.append(["category", "realized_gain", "manual_tax_rate", "calculated_tax"])
    tax.append(["STCG", "=SUM(STCG!G:G)", 0.0, "=MAX(0,B2)*C2"])
    tax.append(["LTCG", "=SUM(LTCG!G:G)", 0.0, "=MAX(0,B3)*C3"])
    tax.append(["TOTAL", "=SUM(B2:B3)", None, "=SUM(D2:D3)"])
    tax.append(["POST_TAX_GAIN", "=B4-D4", None, None])
    _style_table(tax)
    for cell in tax["C"][1:3]:
        cell.number_format = _PERCENT_FORMAT
        cell.fill = PatternFill("solid", fgColor="FFF2CC")
    for row in tax.iter_rows(min_row=2, min_col=2, max_col=4):
        for cell in row:
            if cell.column != 3:
                cell.number_format = _NUMBER_FORMAT

    assumptions = wb.create_sheet("Assumptions")
    assumptions.append(["parameter", "value"])
    assumptions.append(["annual_review_frequency", 1])
    assumptions.append(["ltcg_holding_days", int(ltcg_holding_days)])
    assumptions.append(["tax_applied_after_simulation", 1])
    assumptions.append(["manual_tax_rates", 1])
    if factor_authority:
        assumptions.append(["master_factor_source_sha256", factor_authority.get("sha256")])
        assumptions.append(["master_factor_authoritative", 1])
    _style_table(assumptions)

    charts = wb.create_sheet("Charts")
    equity_chart = LineChart()
    equity_chart.title = "Portfolio vs Benchmark"
    equity_chart.y_axis.title = "Value"
    equity_chart.x_axis.title = "Date"
    equity_chart.add_data(Reference(equity_sheet, min_col=3, min_row=1, max_row=equity_sheet.max_row), titles_from_data=True)
    equity_chart.add_data(Reference(equity_sheet, min_col=5, min_row=1, max_row=equity_sheet.max_row), titles_from_data=True)
    equity_chart.set_categories(Reference(equity_sheet, min_col=1, min_row=2, max_row=equity_sheet.max_row))
    equity_chart.height = 9
    equity_chart.width = 18
    charts.add_chart(equity_chart, "A1")

    drawdown_chart = LineChart()
    drawdown_chart.title = "Portfolio Drawdown"
    drawdown_chart.y_axis.title = "Drawdown"
    drawdown_chart.add_data(Reference(drawdown_sheet, min_col=2, min_row=1, max_row=drawdown_sheet.max_row), titles_from_data=True)
    drawdown_chart.set_categories(Reference(drawdown_sheet, min_col=1, min_row=2, max_row=drawdown_sheet.max_row))
    drawdown_chart.height = 8
    drawdown_chart.width = 18
    charts.add_chart(drawdown_chart, "A20")

    if price_index_sheet.max_column > 1 and price_index_sheet.max_row > 1:
        price_chart = LineChart()
        price_chart.title = "Historical Price Movement (Indexed to 100)"
        price_chart.y_axis.title = "Index"
        price_chart.x_axis.title = "Date"
        price_chart.add_data(
            Reference(
                price_index_sheet,
                min_col=2,
                max_col=price_index_sheet.max_column,
                min_row=1,
                max_row=price_index_sheet.max_row,
            ),
            titles_from_data=True,
        )
        price_chart.set_categories(
            Reference(price_index_sheet, min_col=1, min_row=2, max_row=price_index_sheet.max_row)
        )
        price_chart.height = 9
        price_chart.width = 18
        charts.add_chart(price_chart, "A38")

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(path)
    return path

