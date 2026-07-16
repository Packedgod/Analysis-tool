"""Numeric portfolio-report artifact tests."""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from backtest.portfolio_report import write_portfolio_workbook


def test_workbook_preserves_numeric_history_benchmark_tax_lots_and_charts(tmp_path: Path):
    dates = pd.to_datetime(["2022-01-03", "2022-12-30", "2023-01-02", "2023-12-29"])
    equity = pd.DataFrame(
        {
            "ret": [0.0, 0.10, 0.0, 0.12],
            "equity": [100_000.0, 110_000.0, 110_000.0, 123_200.0],
            "drawdown": [0.0, -0.03, -0.01, -0.05],
            "benchmark_equity": [100_000.0, 108_000.0, 108_000.0, 116_640.0],
            "active_ret": [0.0, 0.02, 0.0, 0.04],
        },
        index=dates,
    )
    equity.index.name = "timestamp"
    positions = pd.DataFrame({"AAA.NS": [0.5] * 4, "BBB.NS": [0.5] * 4}, index=dates)
    positions.index.name = "timestamp"
    trades = pd.DataFrame(
        [
            ["2023-03-01", "AAA.NS", "sell", 100.0, 10.0, "review", 500.0, 120, 5.0],
            ["2023-12-29", "BBB.NS", "sell", 200.0, 5.0, "review", 800.0, 500, 8.0],
        ],
        columns=["timestamp", "code", "side", "price", "qty", "reason", "pnl", "holding_days", "return_pct"],
    )

    output = write_portfolio_workbook(
        tmp_path / "portfolio_report.xlsx",
        equity=equity,
        positions=positions,
        trades=trades,
        metrics={"total_return": 0.232, "max_drawdown": -0.05, "trade_count": 2},
    )

    workbook = load_workbook(output, data_only=False)
    expected = {
        "Summary", "Annual", "Equity", "Holdings", "Transactions", "Benchmark",
        "Drawdown", "STCG", "LTCG", "Manual Tax", "Assumptions", "Charts",
    }
    assert expected.issubset(workbook.sheetnames)
    assert workbook["STCG"]["B2"].value == "AAA.NS"
    assert workbook["LTCG"]["B2"].value == "BBB.NS"
    assert workbook["Manual Tax"]["C2"].value == 0
    assert workbook["Manual Tax"]["C3"].value == 0
    assert workbook["Manual Tax"]["D2"].value == "=MAX(0,B2)*C2"
    assert isinstance(workbook["Annual"]["B2"].value, (int, float))
    assert len(workbook["Charts"]._charts) == 2


def test_private_portfolio_rules_do_not_appear_in_guided_frontend():
    root = Path(__file__).resolve().parents[2]
    frontend = (root / "frontend" / "src" / "pages" / "Home.tsx").read_text(encoding="utf-8")

    for private_term in (
        "Private execution contract",
        "portfolio_report.xlsx",
        "STCG",
        "LTCG",
        "50/200",
        "annual review date",
        "manual_tax_rate",
    ):
        assert private_term not in frontend

